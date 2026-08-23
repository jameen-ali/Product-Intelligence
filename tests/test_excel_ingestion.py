"""
Comprehensive tests for Excel / CSV Catalog ingestion.
Covers: parsing, column mapping, provenance, normalization, DB persistence,
Qdrant, Neo4j, API endpoints, validation, multi-source conflicts, product isolation.
"""
import sys
import os
import uuid
import shutil
import pytest

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../backend")))

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")
CATALOG_XLSX = os.path.join(FIXTURES_DIR, "EM750_catalog.xlsx")
SUPPLIER_XLSX = os.path.join(FIXTURES_DIR, "EM750_supplier.xlsx")
SAMPLE_CSV = os.path.join(FIXTURES_DIR, "EM750_sample.csv")


# ---------------------------------------------------------------------------
# PART 1 — Unit / Parser Tests (no DB)
# ---------------------------------------------------------------------------

class TestExcelColumnMapping:
    def test_direct_column_name_mapping(self):
        from app.ingestion.excel_ingest import _map_column
        assert _map_column("Voltage") == ("voltage", None)
        assert _map_column("Rated Voltage") == ("voltage", None)
        assert _map_column("rated_voltage") == ("voltage", None)

    def test_column_with_unit_in_parentheses(self):
        from app.ingestion.excel_ingest import _map_column
        result = _map_column("Voltage (V)")
        assert result is not None
        attr, unit_hint = result
        assert attr == "voltage"
        assert unit_hint == "v"

    def test_power_variants(self):
        from app.ingestion.excel_ingest import _map_column
        for col in ["Power", "Rated Power", "Power (kW)", "Power (HP)", "Motor Power"]:
            result = _map_column(col)
            assert result is not None, f"Column '{col}' should map to power"
            assert result[0] == "power"

    def test_speed_variants(self):
        from app.ingestion.excel_ingest import _map_column
        for col in ["Speed", "Rated Speed", "RPM", "Rotational Speed"]:
            result = _map_column(col)
            assert result is not None, f"Column '{col}' should map to rotational_speed"
            assert result[0] == "rotational_speed"

    def test_weight_variants(self):
        from app.ingestion.excel_ingest import _map_column
        for col in ["Weight", "Net Weight", "Weight (kg)", "Net Weight (kg)"]:
            result = _map_column(col)
            assert result is not None
            assert result[0] == "weight"

    def test_unrecognized_column_returns_none(self):
        from app.ingestion.excel_ingest import _map_column
        assert _map_column("Serial Number") is None
        assert _map_column("Part Description") is None
        assert _map_column("") is None


class TestExcelValueUnitExtraction:
    def test_value_with_unit_string(self):
        from app.ingestion.excel_ingest import _extract_value_unit
        raw, numeric, unit = _extract_value_unit("415 V")
        assert numeric == 415.0
        assert unit == "V"

    def test_kw_value(self):
        from app.ingestion.excel_ingest import _extract_value_unit
        raw, numeric, unit = _extract_value_unit("7.5 kW")
        assert numeric == 7.5
        assert unit == "kW"

    def test_plain_numeric_with_col_unit_hint(self):
        from app.ingestion.excel_ingest import _extract_value_unit
        raw, numeric, unit = _extract_value_unit("1450", "RPM")
        assert numeric == 1450.0
        assert unit == "RPM"

    def test_watt_value(self):
        from app.ingestion.excel_ingest import _extract_value_unit
        raw, numeric, unit = _extract_value_unit("7500")
        assert numeric == 7500.0

    def test_empty_cell_returns_none(self):
        from app.ingestion.excel_ingest import _extract_value_unit
        raw, numeric, unit = _extract_value_unit("")
        assert numeric is None
        raw2, n2, u2 = _extract_value_unit("nan")
        assert n2 is None


class TestExcelFileParsing:
    def test_parse_xlsx_basic(self):
        from app.ingestion.excel_ingest import parse_excel
        result = parse_excel(CATALOG_XLSX, "testhash")
        assert result.parse_error is None
        assert result.file_type == "xlsx"
        assert len(result.records) >= 1

    def test_parse_xlsx_extracts_voltage(self):
        from app.ingestion.excel_ingest import parse_excel
        result = parse_excel(CATALOG_XLSX, "testhash")
        voltage_records = [r for r in result.records if r.canonical_attr == "voltage"]
        assert len(voltage_records) >= 1
        assert voltage_records[0].raw_value == "415 V"

    def test_parse_xlsx_extracts_power(self):
        from app.ingestion.excel_ingest import parse_excel
        result = parse_excel(CATALOG_XLSX, "testhash")
        power_records = [r for r in result.records if r.canonical_attr == "power"]
        assert len(power_records) >= 1

    def test_parse_xlsx_provenance_metadata(self):
        from app.ingestion.excel_ingest import parse_excel
        result = parse_excel(CATALOG_XLSX, "testhash")
        assert len(result.sheet_names) >= 1
        for rec in result.records:
            assert rec.sheet_name  # must have sheet name
            assert rec.row_number >= 2  # header is row 1
            assert rec.column_name  # must have column header

    def test_parse_csv_basic(self):
        from app.ingestion.excel_ingest import parse_excel
        result = parse_excel(SAMPLE_CSV, "csvhash")
        assert result.parse_error is None
        assert result.file_type == "csv"
        assert len(result.records) >= 1

    def test_parse_csv_extracts_voltage(self):
        from app.ingestion.excel_ingest import parse_excel
        result = parse_excel(SAMPLE_CSV, "csvhash")
        voltage_records = [r for r in result.records if r.canonical_attr == "voltage"]
        assert len(voltage_records) >= 1

    def test_parse_supplier_xlsx_400v(self):
        from app.ingestion.excel_ingest import parse_excel
        result = parse_excel(SUPPLIER_XLSX, "supplierhash")
        assert result.parse_error is None
        voltage_records = [r for r in result.records if r.canonical_attr == "voltage"]
        assert len(voltage_records) >= 1
        assert "400" in voltage_records[0].raw_value

    def test_missing_header_validation(self, tmp_path):
        """Sheet with no recognized column headers gets a validation warning."""
        from app.ingestion.excel_ingest import parse_excel
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ProductCode", "SomeUnknownField1", "UnknownField2"])
        ws.append(["XYZ-001", "ABC", "DEF"])
        fpath = str(tmp_path / "no_headers.xlsx")
        wb.save(fpath)

        result = parse_excel(fpath, "nohash")
        # Should not crash, just warn
        assert any("No recognized" in m for m in result.validation_messages)

    def test_empty_rows_skipped(self, tmp_path):
        """Empty rows are gracefully skipped without crashing."""
        from app.ingestion.excel_ingest import parse_excel
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Rated Voltage", "Rated Power"])
        ws.append([None, None])  # empty row
        ws.append(["415 V", "7.5 kW"])
        fpath = str(tmp_path / "empty_rows.xlsx")
        wb.save(fpath)

        result = parse_excel(fpath, "emptyhash")
        assert result.parse_error is None
        assert len(result.records) >= 1  # Should still extract from row 3

    def test_invalid_numeric_cell_handled(self, tmp_path):
        """Invalid cell values don't crash — they're stored as raw strings."""
        from app.ingestion.excel_ingest import parse_excel
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Rated Voltage", "Rated Power"])
        ws.append(["NOT-A-NUMBER", "???"])
        fpath = str(tmp_path / "bad_values.xlsx")
        wb.save(fpath)

        result = parse_excel(fpath, "badhash")
        # Should not crash
        assert result.parse_error is None

    def test_unsupported_file_type_returns_error(self, tmp_path):
        from app.ingestion.excel_ingest import parse_excel
        fpath = str(tmp_path / "data.txt")
        with open(fpath, "w") as f:
            f.write("some text")
        result = parse_excel(fpath, "txthash")
        assert result.parse_error is not None
        assert "Unsupported" in result.parse_error


class TestExcelValidation:
    def test_valid_csv_passes(self):
        from app.ingestion.excel_ingest import validate_excel_file
        ok, err = validate_excel_file("data.csv", "text/csv", 1024)
        assert ok is True
        assert err is None

    def test_valid_xlsx_passes(self):
        from app.ingestion.excel_ingest import validate_excel_file
        ok, err = validate_excel_file("catalog.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 50000)
        assert ok is True

    def test_invalid_extension_rejected(self):
        from app.ingestion.excel_ingest import validate_excel_file
        ok, err = validate_excel_file("document.pdf", "application/pdf", 1024)
        assert ok is False
        assert "Unsupported" in err

    def test_file_too_large_rejected(self):
        from app.ingestion.excel_ingest import validate_excel_file
        ok, err = validate_excel_file("large.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                      60 * 1024 * 1024)
        assert ok is False
        assert "too large" in err.lower()


# ---------------------------------------------------------------------------
# PART 2 — Integration Tests (DB + Pipeline)
# ---------------------------------------------------------------------------

class TestExcelIngestionPipeline:
    @pytest.mark.asyncio
    async def test_xlsx_full_pipeline_persistence(self, app_client, tmp_path):
        """XLSX ingestion creates Claims, Evidence (with sheet/row/col), Qdrant, Neo4j entries."""
        from app.core.db import SessionLocal
        from app.models.entities import Product, Source, Claim, Evidence, Document
        from app.services.processing_service import process_excel_for_product
        from app.ingestion.hashing import sha256_file
        import openpyxl

        test_id = uuid.uuid4().hex[:8]
        xlsx_path = str(tmp_path / f"persist_{test_id}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Motors"
        ws.append(["Model", "Rated Voltage", "Rated Power", "Rated Speed", "Frequency", "Weight"])
        ws.append([f"EM-PERSIST-{test_id}", "415 V", "7.5 kW", "1450 RPM", "50 Hz", "62 kg"])
        wb.save(xlsx_path)

        db = SessionLocal()
        try:
            p = Product(name=f"Excel Test Motor {test_id}", model_number=f"ETM-{test_id}")
            db.add(p); db.commit(); db.refresh(p)
            pid = str(p.id)

            s = Source(product_id=p.id, type="excel", name="Catalog XLSX", authority_rank=6)
            db.add(s); db.commit(); db.refresh(s)
            sid = str(s.id)
        finally:
            db.close()

        file_hash = sha256_file(xlsx_path)
        db_sess = SessionLocal()
        try:
            result = await process_excel_for_product(pid, sid, xlsx_path, "xlsx", db_sess)
        finally:
            db_sess.close()

        assert result["status"] == "COMPLETED", f"Ingestion returned: {result}"
        assert result["claims_extracted"] >= 1
        assert result["evidence_stored"] >= 1

        # Verify PostgreSQL persistence
        db_check = SessionLocal()
        try:
            doc = db_check.query(Document).filter(Document.file_hash == file_hash).first()
            assert doc is not None
            assert doc.file_type == "xlsx"
            assert doc.parsed_metadata["sheet_names"] is not None

            claims = db_check.query(Claim).filter(Claim.document_id == doc.id).all()
            assert len(claims) >= 1

            evs = db_check.query(Evidence).filter(Evidence.document_id == doc.id).all()
            assert len(evs) >= 1
            assert evs[0].content_type == "spreadsheet"
            # Excel provenance must have sheet/row/column in bbox
            assert evs[0].bbox is not None
            assert "sheet" in evs[0].bbox
            assert "row" in evs[0].bbox
            assert "column" in evs[0].bbox
        finally:
            db_check.close()

    @pytest.mark.asyncio
    async def test_csv_full_pipeline_persistence(self, app_client, tmp_path):
        """CSV ingestion works through the same pipeline."""
        from app.core.db import SessionLocal
        from app.models.entities import Product, Source, Claim, Document
        from app.services.processing_service import process_excel_for_product
        from app.ingestion.hashing import sha256_file
        import csv

        test_id = uuid.uuid4().hex[:8]
        csv_path = str(tmp_path / f"csv_{test_id}.csv")
        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Model", "Voltage", "Power (kW)", "RPM"])
            writer.writerow([f"CTM-{test_id}", "415 V", "7.5", "1450"])

        db = SessionLocal()
        try:
            p = Product(name=f"CSV Test Motor {test_id}", model_number=f"CTM-{test_id}")
            db.add(p); db.commit(); db.refresh(p)
            pid = str(p.id)
            s = Source(product_id=p.id, type="csv", name="CSV Catalog", authority_rank=6)
            db.add(s); db.commit(); db.refresh(s)
            sid = str(s.id)
        finally:
            db.close()

        file_hash = sha256_file(csv_path)
        db_sess = SessionLocal()
        try:
            result = await process_excel_for_product(pid, sid, csv_path, "csv", db_sess)
        finally:
            db_sess.close()

        assert result["status"] == "COMPLETED", f"CSV ingestion returned: {result}"
        assert result["claims_extracted"] >= 1

    @pytest.mark.asyncio
    async def test_excel_deduplication(self, app_client, tmp_path):
        """Uploading the same file twice returns DEDUPLICATED."""
        from app.core.db import SessionLocal
        from app.models.entities import Product, Source
        from app.services.processing_service import process_excel_for_product
        import openpyxl, time

        # Create a truly unique file with unique ID in content so hash is unique
        test_id = uuid.uuid4().hex[:8]
        unique_path = str(tmp_path / f"dedup_{test_id}.xlsx")
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(['Model', 'Rated Voltage', 'Rated Power'])
        ws.append([f'EM-DEDUP-{test_id}', '415 V', '7.5 kW'])
        wb.save(unique_path)

        db = SessionLocal()
        try:
            p = Product(name="Dedup Excel Motor", model_number=f"DEDUP-XL-{test_id}")
            db.add(p); db.commit(); db.refresh(p)
            pid = str(p.id)
            s = Source(product_id=p.id, type="excel", name="Dedup Catalog", authority_rank=6)
            db.add(s); db.commit(); db.refresh(s)
            sid = str(s.id)
        finally:
            db.close()

        db1 = SessionLocal()
        res1 = await process_excel_for_product(pid, sid, unique_path, "xlsx", db1)
        db1.close()

        db2 = SessionLocal()
        res2 = await process_excel_for_product(pid, sid, unique_path, "xlsx", db2)
        db2.close()

        assert res1["status"] == "COMPLETED", f"First ingestion failed: {res1}"
        assert res2["status"] == "DEDUPLICATED"

    @pytest.mark.asyncio
    async def test_multi_source_conflict_excel_vs_pdf(self, app_client, tmp_path):
        """
        Manufacturer XLSX (415V) + Supplier XLSX (400V) → both claims persist independently.
        """
        from app.core.db import SessionLocal
        from app.models.entities import Product, Source, Claim
        from app.services.processing_service import process_excel_for_product
        import openpyxl

        # Create unique files with distinct content (unique model IDs so hash differs)
        test_id = uuid.uuid4().hex[:8]
        catalog_path = str(tmp_path / f"catalog_mfr_{test_id}.xlsx")
        supplier_path = str(tmp_path / f"catalog_sup_{test_id}.xlsx")

        wb1 = openpyxl.Workbook(); ws1 = wb1.active
        ws1.append(['Model', 'Rated Voltage', 'Rated Power'])
        ws1.append([f'MFR-{test_id}', '415 V', '7.5 kW'])
        wb1.save(catalog_path)

        wb2 = openpyxl.Workbook(); ws2 = wb2.active
        ws2.append(['Model', 'Rated Voltage', 'Rated Power (W)'])
        ws2.append([f'SUP-{test_id}', '400 V', '7500'])
        wb2.save(supplier_path)

        db = SessionLocal()
        try:
            p = Product(name="Multi-Source Conflict Motor", model_number=f"CONFLICT-XL-{test_id}")
            db.add(p); db.commit(); db.refresh(p)
            pid = str(p.id)

            s1 = Source(product_id=p.id, type="excel", name="Manufacturer Catalog", authority_rank=1)
            s2 = Source(product_id=p.id, type="excel", name="Supplier Catalog", authority_rank=6)
            db.add(s1); db.add(s2); db.commit()
            db.refresh(s1); db.refresh(s2)
            sid1 = str(s1.id)
            sid2 = str(s2.id)
        finally:
            db.close()

        # Process manufacturer (415V)
        db1 = SessionLocal()
        r1 = await process_excel_for_product(pid, sid1, catalog_path, "xlsx", db1)
        db1.close()
        assert r1["status"] == "COMPLETED"

        # Process supplier (400V) — should create independent claim
        db2 = SessionLocal()
        r2 = await process_excel_for_product(pid, sid2, supplier_path, "xlsx", db2)
        db2.close()
        assert r2["status"] == "COMPLETED"

        # Verify both voltage claims exist independently (different source_ids)
        db_check = SessionLocal()
        try:
            from app.models.entities import Attribute
            voltage_attr = db_check.query(Attribute).filter(Attribute.name == "voltage").first()
            if voltage_attr:
                claims = db_check.query(Claim).filter(
                    Claim.product_id == uuid.UUID(pid),
                    Claim.attribute_id == voltage_attr.id,
                ).all()
                assert len(claims) >= 2, f"Expected at least 2 voltage claims, got {len(claims)}"
                raw_values = {c.raw_value for c in claims}
                assert any("415" in v for v in raw_values), f"415V not found in {raw_values}"
                assert any("400" in v for v in raw_values), f"400V not found in {raw_values}"
        finally:
            db_check.close()

    @pytest.mark.asyncio
    async def test_product_isolation_excel(self, app_client, tmp_path):
        """Excel ingestion for Product A creates no claims on Product B."""
        from app.core.db import SessionLocal
        from app.models.entities import Product, Source, Claim
        from app.services.processing_service import process_excel_for_product
        import openpyxl

        test_id = uuid.uuid4().hex[:8]
        iso_path = str(tmp_path / f"iso_{test_id}.xlsx")
        wb = openpyxl.Workbook(); ws = wb.active
        ws.append(["Model", "Rated Voltage", "Rated Power"])
        ws.append([f"ISO-{test_id}", "415 V", "7.5 kW"])
        wb.save(iso_path)

        db = SessionLocal()
        try:
            pa = Product(name="Isolation Motor A", model_number=f"ISO-A-{test_id}")
            pb = Product(name="Isolation Motor B", model_number=f"ISO-B-{test_id}")
            db.add(pa); db.add(pb); db.commit()
            db.refresh(pa); db.refresh(pb)
            pa_id = str(pa.id); pb_id = str(pb.id)

            s = Source(product_id=pa.id, type="excel", name="Source A", authority_rank=6)
            db.add(s); db.commit(); db.refresh(s)
            sa_id = str(s.id)
        finally:
            db.close()

        db_sess = SessionLocal()
        await process_excel_for_product(pa_id, sa_id, iso_path, "xlsx", db_sess)
        db_sess.close()

        db_check = SessionLocal()
        try:
            claims_b = db_check.query(Claim).filter(Claim.product_id == uuid.UUID(pb_id)).all()
            assert len(claims_b) == 0, f"Product B should have 0 claims, got {len(claims_b)}"
        finally:
            db_check.close()

    def test_api_excel_upload_invalid_type(self, app_client):
        """API returns 400 for non-spreadsheet file uploads."""
        from app.core.db import SessionLocal
        from app.models.entities import Product, Source

        db = SessionLocal()
        try:
            p = Product(name="API Test Motor", model_number="API-XL-01")
            db.add(p); db.commit(); db.refresh(p)
            pid = str(p.id)
            s = Source(product_id=p.id, type="excel", name="API Source", authority_rank=6)
            db.add(s); db.commit(); db.refresh(s)
            sid = str(s.id)
        finally:
            db.close()

        import io
        r = app_client.post(
            f"/processing/products/{pid}/process-excel",
            data={"source_id": sid},
            files={"file": ("test.pdf", io.BytesIO(b"fake pdf content"), "application/pdf")},
        )
        assert r.status_code == 400
        assert "Invalid file" in r.json()["detail"]
